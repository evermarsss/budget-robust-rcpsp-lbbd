# -*- coding: utf-8 -*-
"""
Shared utilities for sequential benchmark runners.

Features:
- Clean terminal dashboard with spinner, progress bars and ETA.
- Polished Excel workbook with first sheet README_摘要.
- Sequential execution only; no multiprocessing, no concurrent Gurobi jobs.
"""
from __future__ import annotations

import math
import os
import shutil
import statistics
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception as exc:
    raise RuntimeError("Please install openpyxl first: pip install openpyxl") from exc

SPINNER = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe(v: Any) -> Any:
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if isinstance(v, (str, int, float, bool)) or v is None:
        return v
    return str(v)


def to_num(v: Any) -> Optional[float]:
    try:
        if v is None or v == "":
            return None
        x = float(v)
        if math.isnan(x) or math.isinf(x):
            return None
        return x
    except Exception:
        return None


def fmt_seconds(sec: Optional[float]) -> str:
    if sec is None:
        return "--"
    sec = max(0.0, float(sec))
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    s = int(sec % 60)
    if h:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def fmt_num(x: Any, digits: int = 3) -> str:
    y = to_num(x)
    if y is None:
        return "--"
    if abs(y) >= 1000:
        return f"{y:,.1f}"
    return f"{y:.{digits}f}"


def fmt_gap(gap: Any) -> str:
    y = to_num(gap)
    if y is None:
        return "--"
    return f"{100*y:.2f}%"


def bar(done: int, total: int, width: int = 28) -> str:
    if total <= 0:
        return "[" + "░" * width + "]  0.0%"
    ratio = min(1.0, max(0.0, done / total))
    filled = int(round(width * ratio))
    return "[" + "█" * filled + "░" * (width - filled) + f"] {100*ratio:5.1f}%"


@dataclass
class RunInfo:
    name: str
    method: str
    dataset: str
    output_path: str
    total_jobs: int
    time_limit: float
    threads: int
    mode: str = "sequential"
    started_at: str = field(default_factory=now_str)
    start_wall: float = field(default_factory=time.time)
    gamma_list: Sequence[Any] = field(default_factory=list)
    alpha_list: Sequence[Any] = field(default_factory=list)
    group_range: str = ""
    sample_indices: Sequence[int] = field(default_factory=list)


class TerminalDashboard:
    def __init__(self, run_info: RunInfo, refresh_interval: float = 1.0, enabled: bool = True):
        self.run_info = run_info
        self.refresh_interval = refresh_interval
        self.enabled = enabled
        self.last_draw = 0.0
        self.frame = 0
        self.recent: List[Dict[str, Any]] = []
        self.hard: List[Dict[str, Any]] = []
        self.completed_runtimes: List[float] = []
        self.last_text_len = 0

    def add_completed(self, result: Dict[str, Any]) -> None:
        self.recent.insert(0, result)
        self.recent = self.recent[:8]
        rt = to_num(result.get("runtime_total")) or to_num(result.get("total_time"))
        if rt is not None:
            self.completed_runtimes.append(rt)
            self.hard.append(result)
            self.hard = sorted(
                self.hard,
                key=lambda r: (to_num(r.get("runtime_total")) or to_num(r.get("total_time")) or 0.0),
                reverse=True,
            )[:5]

    def eta_values(self, done_jobs: int) -> Tuple[Optional[float], Optional[float], Optional[float]]:
        pending = max(0, self.run_info.total_jobs - done_jobs)
        if pending == 0:
            return 0.0, 0.0, 0.0
        if not self.completed_runtimes:
            return None, None, pending * self.run_info.time_limit
        median_rt = statistics.median(self.completed_runtimes)
        mean_rt = statistics.mean(self.completed_runtimes)
        return pending * median_rt, pending * mean_rt, pending * self.run_info.time_limit

    def draw(self, done_jobs: int, current_index: int, current: Optional[Dict[str, Any]], force: bool = False) -> None:
        if not self.enabled:
            return
        now = time.time()
        if not force and now - self.last_draw < self.refresh_interval:
            return
        self.last_draw = now
        spin = SPINNER[self.frame % len(SPINNER)]
        self.frame += 1
        elapsed = now - self.run_info.start_wall
        eta_med, eta_mean, eta_worst = self.eta_values(done_jobs)
        opt = sum(1 for r in self.recent + self.hard if str(r.get("status_name")) == "OPTIMAL")  # local only, final Excel is source of truth
        lines: List[str] = []
        lines.append(f"{spin} Robust RCPSP Benchmark Runner   [{self.run_info.method}]   mode={self.run_info.mode}   Threads={self.run_info.threads}")
        lines.append(f"Dataset={self.run_info.dataset} | Gamma={list(self.run_info.gamma_list)} | alpha={list(self.run_info.alpha_list)} | samples={list(self.run_info.sample_indices)} | groups={self.run_info.group_range}")
        lines.append(f"Output: {self.run_info.output_path}")
        lines.append("-" * 110)
        lines.append(f"Overall progress  {bar(done_jobs, self.run_info.total_jobs)}   {done_jobs}/{self.run_info.total_jobs}")
        lines.append(f"Elapsed={fmt_seconds(elapsed)} | ETA median={fmt_seconds(eta_med)} | ETA mean={fmt_seconds(eta_mean)} | ETA worst={fmt_seconds(eta_worst)}")
        lines.append("-" * 110)
        if current:
            cur_elapsed = to_num(current.get("elapsed")) or 0.0
            tl = to_num(current.get("time_limit")) or self.run_info.time_limit
            lines.append(f"Current job {current_index}/{self.run_info.total_jobs}  {Path(str(current.get('file',''))).name}")
            lines.append(f"  method={current.get('method', self.run_info.method)}  G={current.get('gamma')}  alpha={current.get('alpha')}  TL={tl}s")
            lines.append(f"  runtime {bar(int(cur_elapsed), int(tl), width=28)}   {fmt_seconds(cur_elapsed)} / {fmt_seconds(tl)}")
            lines.append(f"  obj={fmt_num(current.get('objective'))}  bound={fmt_num(current.get('best_bound'))}  gap={fmt_gap(current.get('gap'))}  nodes={fmt_num(current.get('nodes'),0)}")
            # LBBD-specific fields are harmless for Gurobi if missing.
            lines.append(f"  SP={current.get('sp_calls','--')}  MFS={current.get('mfs_cuts','--')}  path_user={current.get('path_user','--')}  path_cuts={current.get('path_cuts','--')}  sol={current.get('sol_count','--')}")
        else:
            lines.append("Current job: --")
        lines.append("-" * 110)
        lines.append("Recent completed:")
        if not self.recent:
            lines.append("  --")
        for r in self.recent[:6]:
            rt = to_num(r.get("runtime_total")) or to_num(r.get("total_time"))
            lines.append(
                f"  {Path(str(r.get('file',''))).name:<12} G={r.get('gamma')} a={r.get('alpha')} "
                f"{str(r.get('status_name')):<11} time={fmt_seconds(rt):>8} gap={fmt_gap(r.get('gap')):>8} obj={fmt_num(r.get('objective'))}"
            )
        lines.append("-" * 110)
        lines.append("Slowest finished so far:")
        if not self.hard:
            lines.append("  --")
        for r in self.hard[:5]:
            rt = to_num(r.get("runtime_total")) or to_num(r.get("total_time"))
            lines.append(
                f"  {Path(str(r.get('file',''))).name:<12} G={r.get('gamma')} a={r.get('alpha')} "
                f"{str(r.get('status_name')):<11} time={fmt_seconds(rt):>8} gap={fmt_gap(r.get('gap')):>8}"
            )
        text = "\n".join(lines)
        # Redraw as a fixed dashboard. On some Windows/IDE terminals ANSI home/clear
        # is not honored, which causes the dashboard to scroll forever. Use native
        # `cls` on Windows so the screen is cleared before every refresh.
        if os.name == "nt":
            os.system("cls")
            sys.stdout.write(text + "\n")
        else:
            sys.stdout.write("\033[2J\033[H" + text + "\n")
        sys.stdout.flush()


class ExcelReport:
    def __init__(self, output_path: str, run_info: RunInfo, headers: List[str], summary_keys: List[str], config: Dict[str, Any]):
        self.output_path = output_path
        self.run_info = run_info
        self.headers = headers
        self.summary_keys = summary_keys
        self.config = config
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "README_摘要"
        self.ws_readme = ws
        self.ws_all = self.wb.create_sheet("all_results")
        self.ws_summary = self.wb.create_sheet("summary_by_param")
        self.ws_status = self.wb.create_sheet("status_summary")
        self.ws_abnormal = self.wb.create_sheet("abnormal")
        self.ws_hard = self.wb.create_sheet("hard_cases")
        self.ws_config = self.wb.create_sheet("config")
        self.result_rows: Dict[int, int] = {}
        self.abnormal_run_ids: set[int] = set()
        self.last_save = 0.0
        self._init_table(self.ws_all, headers)
        self._init_table(self.ws_abnormal, headers)
        self._init_table(self.ws_hard, headers)
        self._init_table(self.ws_summary, summary_keys + ["runs", "optimal", "time_limit", "error", "abnormal", "avg_runtime", "median_runtime", "avg_gap", "max_gap", "avg_objective"])
        self._init_table(self.ws_status, ["status_name", "runs", "avg_runtime", "avg_gap"])
        self._init_table(self.ws_config, ["key", "value"])
        for k, v in config.items():
            self.ws_config.append([k, str(v)])
        self.rebuild_all_summaries()
        self.save(force=True)

    def _init_table(self, ws, headers: List[str]) -> None:
        ws.append(headers)
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.freeze_panes = "A2"
        for idx, h in enumerate(headers, start=1):
            width = min(max(len(str(h)) + 2, 10), 28)
            if h in ("file", "path", "error"):
                width = 38
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _collect_rows(self) -> List[Dict[str, Any]]:
        rows = []
        for row in self.ws_all.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            rows.append({h: row[i] if i < len(row) else None for i, h in enumerate(self.headers)})
        return rows

    def add_or_update(self, result: Dict[str, Any], force_save: bool = False) -> None:
        run_id = int(result.get("run_id", len(self.result_rows) + 1))
        result["run_id"] = run_id
        if run_id not in self.result_rows:
            row = self.ws_all.max_row + 1
            self.ws_all.append([None] * len(self.headers))
            self.result_rows[run_id] = row
        row = self.result_rows[run_id]
        for col, h in enumerate(self.headers, start=1):
            self.ws_all.cell(row=row, column=col, value=safe(result.get(h)))
        self._style_row(self.ws_all, row, result)
        if result.get("abnormal") and run_id not in self.abnormal_run_ids:
            self.ws_abnormal.append([safe(result.get(h)) for h in self.headers])
            self._style_row(self.ws_abnormal, self.ws_abnormal.max_row, result)
            self.abnormal_run_ids.add(run_id)
        if force_save:
            self.rebuild_all_summaries()
        self.save(force=force_save)

    def _style_row(self, ws, row: int, result: Dict[str, Any]) -> None:
        status = str(result.get("status_name", ""))
        abnormal = bool(result.get("abnormal", False))
        color = "E2F0D9"
        if status == "RUNNING":
            color = "D9EAF7"
        elif abnormal or status in ("TIME_LIMIT", "ERROR", "INFEASIBLE", "INF_OR_UNBD"):
            color = "FFF2CC" if status == "TIME_LIMIT" else "F4CCCC"
        fill = PatternFill("solid", fgColor=color)
        for cell in ws[row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    def rebuild_all_summaries(self) -> None:
        rows = self._collect_rows()
        self._rebuild_readme(rows)
        self._rebuild_summary(rows)
        self._rebuild_status(rows)
        self._rebuild_hard(rows)

    def _clear_keep_header(self, ws) -> None:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    def _rebuild_summary(self, rows: List[Dict[str, Any]]) -> None:
        self._clear_keep_header(self.ws_summary)
        groups: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
        for r in rows:
            key = tuple(r.get(k) for k in self.summary_keys)
            groups[key].append(r)
        for key, items in sorted(groups.items(), key=lambda kv: str(kv[0])):
            runs = len(items)
            optimal = sum(str(r.get("status_name")) == "OPTIMAL" for r in items)
            tl = sum(str(r.get("status_name")) == "TIME_LIMIT" for r in items)
            err = sum(str(r.get("status_name")) == "ERROR" for r in items)
            abnormal = sum(bool(r.get("abnormal")) for r in items)
            runtimes = [to_num(r.get("runtime_total")) or to_num(r.get("total_time")) for r in items]
            runtimes = [x for x in runtimes if x is not None]
            gaps = [to_num(r.get("gap")) for r in items]
            gaps = [x for x in gaps if x is not None]
            objs = [to_num(r.get("objective")) for r in items]
            objs = [x for x in objs if x is not None]
            self.ws_summary.append(list(key) + [
                runs, optimal, tl, err, abnormal,
                statistics.mean(runtimes) if runtimes else None,
                statistics.median(runtimes) if runtimes else None,
                statistics.mean(gaps) if gaps else None,
                max(gaps) if gaps else None,
                statistics.mean(objs) if objs else None,
            ])

    def _rebuild_status(self, rows: List[Dict[str, Any]]) -> None:
        self._clear_keep_header(self.ws_status)
        groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for r in rows:
            groups[str(r.get("status_name"))].append(r)
        for status, items in sorted(groups.items()):
            rts = [to_num(r.get("runtime_total")) or to_num(r.get("total_time")) for r in items]
            rts = [x for x in rts if x is not None]
            gaps = [to_num(r.get("gap")) for r in items]
            gaps = [x for x in gaps if x is not None]
            self.ws_status.append([status, len(items), statistics.mean(rts) if rts else None, statistics.mean(gaps) if gaps else None])

    def _rebuild_hard(self, rows: List[Dict[str, Any]]) -> None:
        self._clear_keep_header(self.ws_hard)
        ranked = sorted(rows, key=lambda r: (to_num(r.get("runtime_total")) or to_num(r.get("total_time")) or 0.0), reverse=True)[:30]
        for r in ranked:
            self.ws_hard.append([safe(r.get(h)) for h in self.headers])
            self._style_row(self.ws_hard, self.ws_hard.max_row, r)

    def _rebuild_readme(self, rows: List[Dict[str, Any]]) -> None:
        ws = self.ws_readme
        ws.delete_rows(1, ws.max_row)
        title_fill = PatternFill("solid", fgColor="1F4E78")
        title_font = Font(color="FFFFFF", bold=True, size=14)
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        bold = Font(bold=True)
        ws.append(["Robust RCPSP Benchmark 摘要"])
        ws.merge_cells("A1:E1")
        ws["A1"].fill = title_fill
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center")
        total = len(rows)
        statuses = defaultdict(int)
        gammas = set()
        alphas = set()
        methods = set()
        datasets = set()
        runtime_sum = 0.0
        for r in rows:
            statuses[str(r.get("status_name"))] += 1
            if r.get("gamma") is not None: gammas.add(r.get("gamma"))
            if r.get("alpha") is not None: alphas.add(r.get("alpha"))
            if r.get("method") is not None: methods.add(r.get("method"))
            if r.get("dataset") is not None: datasets.add(r.get("dataset"))
            runtime_sum += to_num(r.get("runtime_total")) or to_num(r.get("total_time")) or 0.0
        elapsed = time.time() - self.run_info.start_wall
        lines = [
            ("程序名称", self.run_info.name),
            ("方法", self.run_info.method),
            ("运行模式", self.run_info.mode),
            ("创建时间", self.run_info.started_at),
            ("最后保存", now_str()),
            ("当前程序已运行", fmt_seconds(elapsed)),
            ("Excel中已记录数据条数", total),
            ("已完成求解时间合计", fmt_seconds(runtime_sum)),
            ("数据集", ", ".join(map(str, sorted(datasets))) or self.run_info.dataset),
            ("Gamma 个数/取值", f"{len(gammas)} / {sorted(gammas, key=str)}"),
            ("alpha 个数/取值", f"{len(alphas)} / {sorted(alphas, key=str)}"),
            ("方法个数/取值", f"{len(methods)} / {sorted(methods, key=str)}"),
            ("总任务数", self.run_info.total_jobs),
            ("当前完成比例", f"{total}/{self.run_info.total_jobs}"),
            ("Threads", self.run_info.threads),
            ("TimeLimit", self.run_info.time_limit),
            ("输出文件", self.output_path),
        ]
        ws.append([])
        ws.append(["基本信息", "值"])
        ws["A3"].fill = section_fill; ws["B3"].fill = section_fill; ws["A3"].font = bold; ws["B3"].font = bold
        for k, v in lines:
            ws.append([k, v])
        ws.append([])
        ws.append(["状态统计", "数量"])
        row0 = ws.max_row
        ws.cell(row0,1).fill = section_fill; ws.cell(row0,2).fill = section_fill; ws.cell(row0,1).font = bold; ws.cell(row0,2).font = bold
        for k, v in sorted(statuses.items()):
            ws.append([k, v])
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["C"].width = 20
        ws.freeze_panes = "A3"

    def save(self, force: bool = False) -> None:
        now = time.time()
        if force or now - self.last_save >= 10.0:
            os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
            self.wb.save(self.output_path)
            self.last_save = now

# =============================================================================
# External popup dashboard support (v3)
# =============================================================================
import json
import subprocess


def _atomic_write_json(path: str, payload: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    os.replace(tmp, path)


class ExternalPopupDashboard(TerminalDashboard):
    """Dashboard backend that writes a JSON status file for benchmark_monitor_gui.py.

    It has the same add_completed/draw API as TerminalDashboard, but it does not
    redraw the terminal. The external Tkinter monitor reads the JSON file and
    updates a stable window every few hundred milliseconds, so there is no
    terminal flicker or eye-straining jumping.
    """

    def __init__(self, run_info: RunInfo, status_path: str, refresh_interval: float = 0.3,
                 enabled: bool = True, launch_popup: bool = True):
        super().__init__(run_info, refresh_interval=refresh_interval, enabled=enabled)
        self.status_path = status_path
        self.launch_popup = launch_popup
        self.popup_proc = None
        if enabled and launch_popup:
            self._launch_monitor()

    def _launch_monitor(self) -> None:
        try:
            script = os.path.join(os.path.dirname(__file__), "benchmark_monitor_gui.py")
            if os.path.exists(script):
                self.popup_proc = subprocess.Popen(
                    [sys.executable, script, self.status_path],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0,
                )
        except Exception:
            self.popup_proc = None

    def draw(self, done_jobs: int, current_index: int, current: Optional[Dict[str, Any]], force: bool = False) -> None:
        if not self.enabled:
            return
        now = time.time()
        if not force and now - self.last_draw < self.refresh_interval:
            return
        self.last_draw = now
        self.frame += 1
        elapsed = now - self.run_info.start_wall
        eta_med, eta_mean, eta_worst = self.eta_values(done_jobs)
        payload = {
            "kind": "robust_rcpsp_benchmark_status",
            "version": 3,
            "timestamp": now_str(),
            "frame": self.frame,
            "spinner": SPINNER[self.frame % len(SPINNER)],
            "run_info": {
                "name": self.run_info.name,
                "method": self.run_info.method,
                "dataset": self.run_info.dataset,
                "output_path": self.run_info.output_path,
                "total_jobs": self.run_info.total_jobs,
                "time_limit": self.run_info.time_limit,
                "threads": self.run_info.threads,
                "mode": self.run_info.mode,
                "started_at": self.run_info.started_at,
                "gamma_list": list(self.run_info.gamma_list),
                "alpha_list": list(self.run_info.alpha_list),
                "group_range": self.run_info.group_range,
                "sample_indices": list(self.run_info.sample_indices),
            },
            "progress": {
                "done_jobs": done_jobs,
                "current_index": current_index,
                "total_jobs": self.run_info.total_jobs,
                "elapsed": elapsed,
                "eta_median": eta_med,
                "eta_mean": eta_mean,
                "eta_worst": eta_worst,
            },
            "current": current or {},
            "recent": self.recent[:10],
            "hard": self.hard[:8],
        }
        try:
            _atomic_write_json(self.status_path, payload)
        except Exception:
            pass


def create_dashboard(run_info: RunInfo, mode: str = "popup", refresh_interval: float = 0.3,
                     enabled: bool = True, status_path: str = "", launch_popup: bool = True):
    mode = (mode or "terminal").lower().strip()
    if not enabled:
        return TerminalDashboard(run_info, refresh_interval=refresh_interval, enabled=False)
    if mode in ("popup", "external", "gui", "window"):
        if not status_path:
            base = os.path.splitext(run_info.output_path)[0]
            status_path = base + "_live_status.json"
        return ExternalPopupDashboard(run_info, status_path, refresh_interval=refresh_interval, enabled=True, launch_popup=launch_popup)
    return TerminalDashboard(run_info, refresh_interval=refresh_interval, enabled=True)
