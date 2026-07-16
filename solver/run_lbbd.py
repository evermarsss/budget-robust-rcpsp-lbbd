# -*- coding: utf-8 -*-
"""
Batch benchmark runner for the robust RCPSP LBBD-main algorithm.

Purpose
-------
Run lbbd_main.py on a PSPLIB
J30/J60/J90 style instance library, e.g. j301_1.sm ... j3048_10.sm, with
selectable Gamma/alpha/time-limit parameters.

Main features
-------------
1. Same library-style configuration as the previous Gurobi benchmark scripts.
2. Runs groups j301--j3048 with selected sample indices, Gamma list, Alpha list.
3. Saves results immediately to a NEW Excel workbook.
4. During each solve, prints a one-line live progress display every second:
   current instance runtime, incumbent, bound, gap, nodes, SP/MFS/path-cut stats.
5. Also writes RUNNING progress into Excel at a configurable interval, so if the
   run is interrupted the workbook still contains the current/finished rows.

Dependencies
------------
- gurobipy
- psplib
- openpyxl
- lbbd_main.py in the same folder

Notes
-----
This program runs the proposed LBBD-main algorithm. The direct LG-baseline is
run separately by run_lg.py.
"""
from __future__ import annotations

import copy
import math
import os
import sys
import time
import shutil
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except Exception as exc:  # pragma: no cover
    raise RuntimeError("This benchmark script needs openpyxl. Please run: pip install openpyxl") from exc

import gurobipy as gp
from gurobipy import GRB

from benchmark_ui_common import RunInfo, create_dashboard

# Import the proposed algorithm implementation. Keep this file in the same folder.
import lbbd_main as algorithm

# =============================================================================
# USER CONFIG: library selection
# =============================================================================
DATASET_NAME = "J60"
DATASET_ROOT = r"C:\Users\17717\Downloads\j60.sm"

# PSPLIB J30 naming convention: group 1 -> j301_1.sm, group 48 -> j3048_1.sm.
GROUP_START = 1
GROUP_END = 48
SAMPLE_INDICES = [1]          # use [1] for 48 representative instances; use list(range(1, 11)) for all 480

# Parameters. These can be multiple values; the script runs all combinations.
GAMMA_LIST = [30]
ALPHA_LIST = [0.5]
TIME_LIMIT_LIST = [600]

# Output workbook. If OUTPUT_XLSX is empty, a timestamped file is created.
OUTPUT_DIR = r"C:\Users\17717\Documents\GitHub\RCPSP\benchmark_results"
OUTPUT_XLSX = ""  # e.g. r"C:\...\lbbd_results.xlsx"; leave blank for timestamped new file

# Run control.
STOP_AFTER_FIRST_ERROR = False
SKIP_MISSING_FILES = True
VERIFY_INCUMBENT = True
ABNORMAL_GAP_TOL = 1e-4

# =============================================================================
# USER CONFIG: LBBD-main algorithm switches
# =============================================================================
# These are copied into the algorithm module before every run.
ALGORITHM_CONFIG = {
    "GUROBI_OUTPUT": 0,
    "THREADS": 4,                 # keep Gurobi log off; this wrapper gives clean live output
    "ADD_TRANSITIVITY": False,
    "ADD_PAIRWISE_CONFLICTS": True,
    "USE_INITIAL_SGS_UB": True,
    "SET_MIP_START_FROM_SGS": True,
    "ADD_ETA_CUTOFF_FROM_SGS": True,
    "CACHE_SP_RESULTS": True,
    "MAX_LAZY_PER_MIPSOL": 20,
    "MIP_FOCUS": 2,
    "HEURISTICS_LEVEL": 0.10,
    "NOREL_HEUR_TIME": 5,
    "ADD_PATH_OPT_CUTS": True,
    "PATH_CUT_FROM_SP1_FEASIBLE": True,
    "PATH_CUT_ADD_LAZY_IF_VIOLATED_AT_MIPSOL": True,
    "PATH_CUT_CHECK_MIPNODE": True,
    "PATH_CUT_MIPNODE_FREQ": 1,
    "PATH_CUT_POOL_MAX_SIZE": 2000,
    "PATH_CUT_POOL_SCAN_LIMIT": 200,
    "MAX_PATH_USER_CUTS_PER_NODE": 10,
    "PATH_CUT_VIOL_TOL": 1e-5,
    "PATH_CUT_MIN_COEFF": 1e-8,
    "ADD_MFS_CUTS": True,
    "USE_FULL_NOGOOD_FALLBACK": True,
    "PROGRESS_OUTPUT": False,           # wrapper handles progress output
}

# =============================================================================
# USER CONFIG: live output and Excel saving
# =============================================================================
# v3 live UI. popup = external Tkinter window; terminal = old one-line output; off = no live UI.
UI_MODE = "popup"
LAUNCH_POPUP_UI = True
POPUP_REFRESH_INTERVAL_SEC = 0.5
LIVE_PRINT_INTERVAL_SEC = 0.5
# Live progress mode. "overwrite" keeps exactly one changing line in the terminal.
# Use "newline" only for debugging.
LIVE_PROGRESS_MODE = "overwrite"
LIVE_PROGRESS_MAX_WIDTH = 0  # 0 = auto terminal width; positive int = hard cap
CLEAR_LIVE_LINE_AFTER_SOLVE = True
EXCEL_SAVE_INTERVAL_SEC = 3.0          # set to 1.0 for most aggressive real-time saving
SAVE_EXCEL_DURING_SOLVE = True
PRINT_FINAL_ONE_LINE = True

# =============================================================================
# Internal constants
# =============================================================================
SPINNER_FRAMES = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]

STATUS_NAME = {
    GRB.OPTIMAL: "OPTIMAL",
    GRB.INFEASIBLE: "INFEASIBLE",
    GRB.INF_OR_UNBD: "INF_OR_UNBD",
    GRB.UNBOUNDED: "UNBOUNDED",
    GRB.CUTOFF: "CUTOFF",
    GRB.ITERATION_LIMIT: "ITERATION_LIMIT",
    GRB.NODE_LIMIT: "NODE_LIMIT",
    GRB.TIME_LIMIT: "TIME_LIMIT",
    GRB.SOLUTION_LIMIT: "SOLUTION_LIMIT",
    GRB.INTERRUPTED: "INTERRUPTED",
    GRB.NUMERIC: "NUMERIC",
    GRB.SUBOPTIMAL: "SUBOPTIMAL",
    GRB.INPROGRESS: "INPROGRESS",
    GRB.USER_OBJ_LIMIT: "USER_OBJ_LIMIT",
}

RESULT_HEADERS = [
    "run_id", "dataset", "group", "sample_index", "gamma", "alpha", "time_limit",
    "file", "status", "status_name", "abnormal", "error",
    "objective", "best_bound", "gap", "runtime_solver", "runtime_total", "build_time",
    "nodes", "sol_count", "vars", "constrs", "pair_cuts",
    "initial_lb", "initial_ub", "initial_rule",
    "sp_calls", "sp_infeasible", "sp_feasible", "mfs_cuts", "mfs_repeat", "mfs_miss",
    "mfs_avg_size", "nogood_cuts", "cache_size", "cache_hits",
    "path_pool", "path_user", "path_lazy", "path_cert", "path_cert_fail",
    "verify_trob", "verify_resource_feasible",
    "start_time", "end_time", "last_update",
]

SUMMARY_HEADERS = [
    "dataset", "gamma", "alpha", "time_limit", "runs", "solved_optimal", "abnormal",
    "avg_runtime_total", "avg_gap", "best_obj_avg", "total_sp", "total_mfs", "total_path_user",
]

CONFIG_HEADERS = ["key", "value"]


@dataclass
class RunCase:
    run_id: int
    dataset: str
    group: int
    sample_index: int
    gamma: int
    alpha: float
    time_limit: float
    file_path: str


class ExcelLiveWriter:
    """Small robust Excel writer for real-time benchmark logging."""

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        # First sheet: human-readable summary for quick inspection.
        ws = self.wb.active
        ws.title = "README_摘要"
        self.ws_readme = ws
        self.ws_results = self.wb.create_sheet("all_results")
        self.ws_abnormal = self.wb.create_sheet("abnormal")
        self.ws_summary = self.wb.create_sheet("summary")
        self.ws_config = self.wb.create_sheet("config")
        self._init_sheet(self.ws_results, RESULT_HEADERS)
        self._init_sheet(self.ws_abnormal, RESULT_HEADERS)
        self._init_sheet(self.ws_summary, SUMMARY_HEADERS)
        self._init_sheet(self.ws_config, CONFIG_HEADERS)
        self.result_row_by_run_id: Dict[int, int] = {}
        self.abnormal_written: set[int] = set()
        self.last_save_time = 0.0
        self._write_config()
        self.save(force=True)

    @staticmethod
    def _init_sheet(ws, headers: List[str]) -> None:
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.freeze_panes = "A2"
        widths = {
            "A": 8, "B": 9, "C": 8, "D": 12, "E": 8, "F": 8, "G": 10,
            "H": 38, "I": 10, "J": 14, "K": 10, "L": 30,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def _write_config(self) -> None:
        cfg = {
            "created_at": now_str(),
            "dataset_name": DATASET_NAME,
            "dataset_root": DATASET_ROOT,
            "group_start": GROUP_START,
            "group_end": GROUP_END,
            "sample_indices": str(SAMPLE_INDICES),
            "gamma_list": str(GAMMA_LIST),
            "alpha_list": str(ALPHA_LIST),
            "time_limit_list": str(TIME_LIMIT_LIST),
            "verify_incumbent": VERIFY_INCUMBENT,
            "threads": ALGORITHM_CONFIG.get("THREADS", 0),
            "algorithm_config": str(ALGORITHM_CONFIG),
        }
        for k, v in cfg.items():
            self.ws_config.append([k, str(v)])
        self.ws_config.column_dimensions["A"].width = 28
        self.ws_config.column_dimensions["B"].width = 120

    def add_or_update_result(self, result: Dict[str, Any], force_save: bool = False) -> None:
        run_id = int(result["run_id"])
        if run_id not in self.result_row_by_run_id:
            row = self.ws_results.max_row + 1
            self.ws_results.append([None] * len(RESULT_HEADERS))
            self.result_row_by_run_id[run_id] = row
        row = self.result_row_by_run_id[run_id]
        for col, key in enumerate(RESULT_HEADERS, start=1):
            self.ws_results.cell(row=row, column=col, value=safe_excel_value(result.get(key)))
        self._style_status_row(self.ws_results, row, result.get("abnormal"))
        if force_save:
            self.save(force=True)
        else:
            self.save(force=False)

    def write_abnormal_if_needed(self, result: Dict[str, Any]) -> None:
        run_id = int(result["run_id"])
        if not result.get("abnormal") or run_id in self.abnormal_written:
            return
        self.ws_abnormal.append([safe_excel_value(result.get(k)) for k in RESULT_HEADERS])
        self._style_status_row(self.ws_abnormal, self.ws_abnormal.max_row, True)
        self.abnormal_written.add(run_id)
        self.save(force=True)

    def rebuild_summary(self) -> None:
        # Clear existing summary except header.
        if self.ws_summary.max_row > 1:
            self.ws_summary.delete_rows(2, self.ws_summary.max_row - 1)
        rows = []
        for row in self.ws_results.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            d = {h: row[i] for i, h in enumerate(RESULT_HEADERS)}
            rows.append(d)
        groups: Dict[Tuple[Any, Any, Any, Any], List[Dict[str, Any]]] = {}
        for d in rows:
            key = (d["dataset"], d["gamma"], d["alpha"], d["time_limit"])
            groups.setdefault(key, []).append(d)
        for key, items in sorted(groups.items(), key=lambda kv: str(kv[0])):
            runs = len(items)
            solved = sum(1 for x in items if x.get("status_name") == "OPTIMAL")
            abnormal = sum(1 for x in items if bool(x.get("abnormal")))
            avg_rt = avg([num(x.get("runtime_total")) for x in items])
            avg_gap = avg([num(x.get("gap")) for x in items if num(x.get("gap")) is not None])
            avg_obj = avg([num(x.get("objective")) for x in items if num(x.get("objective")) is not None])
            total_sp = sum(int(num(x.get("sp_calls")) or 0) for x in items)
            total_mfs = sum(int(num(x.get("mfs_cuts")) or 0) for x in items)
            total_path_user = sum(int(num(x.get("path_user")) or 0) for x in items)
            self.ws_summary.append([
                key[0], key[1], key[2], key[3], runs, solved, abnormal,
                avg_rt, avg_gap, avg_obj, total_sp, total_mfs, total_path_user,
            ])
        self.save(force=True)

    def _style_status_row(self, ws, row: int, abnormal: Any) -> None:
        fill = PatternFill("solid", fgColor="FFF2CC") if abnormal else PatternFill("solid", fgColor="E2F0D9")
        if ws.cell(row=row, column=10).value == "RUNNING":
            fill = PatternFill("solid", fgColor="D9EAF7")
        if ws.cell(row=row, column=10).value in ("ERROR", "INFEASIBLE", "INF_OR_UNBD"):
            fill = PatternFill("solid", fgColor="F4CCCC")
        for cell in ws[row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)


    def _rows_as_dicts(self) -> List[Dict[str, Any]]:
        rows = []
        for row in self.ws_results.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            rows.append({h: row[i] if i < len(row) else None for i, h in enumerate(RESULT_HEADERS)})
        return rows

    def rebuild_readme(self) -> None:
        ws = self.ws_readme
        ws.delete_rows(1, ws.max_row)
        title_fill = PatternFill("solid", fgColor="1F4E78")
        title_font = Font(color="FFFFFF", bold=True, size=14)
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        bold = Font(bold=True)
        rows = self._rows_as_dicts()
        statuses = {}
        gammas = set(); alphas = set(); tls = set()
        runtime_sum = 0.0
        for r in rows:
            statuses[str(r.get("status_name"))] = statuses.get(str(r.get("status_name")), 0) + 1
            if r.get("gamma") is not None: gammas.add(r.get("gamma"))
            if r.get("alpha") is not None: alphas.add(r.get("alpha"))
            if r.get("time_limit") is not None: tls.add(r.get("time_limit"))
            runtime_sum += num(r.get("runtime_total")) or 0.0
        ws.append(["Robust RCPSP LBBD-main Benchmark 摘要"])
        ws.merge_cells("A1:E1")
        ws["A1"].fill = title_fill; ws["A1"].font = title_font; ws["A1"].alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append(["基本信息", "值"])
        ws["A3"].fill = section_fill; ws["B3"].fill = section_fill; ws["A3"].font = bold; ws["B3"].font = bold
        info = [
            ("程序", "run_lbbd.py"),
            ("算法", "LBBD-main"),
            ("数据集", DATASET_NAME),
            ("数据路径", DATASET_ROOT),
            ("group范围", f"{GROUP_START}..{GROUP_END}"),
            ("sample_indices", str(SAMPLE_INDICES)),
            ("Gamma个数/取值", f"{len(gammas)} / {sorted(gammas, key=str)}"),
            ("alpha个数/取值", f"{len(alphas)} / {sorted(alphas, key=str)}"),
            ("TimeLimit个数/取值", f"{len(tls)} / {sorted(tls, key=str)}"),
            ("Excel记录条数", len(rows)),
            ("已完成求解时间合计", f"{runtime_sum:.2f}s"),
            ("最后保存", now_str()),
            ("输出文件", self.output_path),
        ]
        for k, v in info:
            ws.append([k, v])
        ws.append([])
        ws.append(["状态统计", "数量"])
        rr = ws.max_row
        ws.cell(rr,1).fill = section_fill; ws.cell(rr,2).fill = section_fill; ws.cell(rr,1).font = bold; ws.cell(rr,2).font = bold
        for k, v in sorted(statuses.items()):
            ws.append([k, v])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 90

    def save(self, force: bool = False) -> None:
        now = time.time()
        if force or now - self.last_save_time >= EXCEL_SAVE_INTERVAL_SEC:
            os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
            try:
                self.rebuild_readme()
            except Exception:
                pass
            self.wb.save(self.output_path)
            self.last_save_time = now


def safe_excel_value(v: Any) -> Any:
    if isinstance(v, (int, float, str, bool)) or v is None:
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v
    return str(v)


def num(v: Any) -> Optional[float]:
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def avg(vals: Iterable[Optional[float]]) -> Optional[float]:
    xs = [v for v in vals if v is not None]
    return sum(xs) / len(xs) if xs else None


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def make_output_path() -> str:
    if OUTPUT_XLSX.strip():
        return OUTPUT_XLSX
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    sample_tag = "s" + "-".join(str(x) for x in SAMPLE_INDICES)
    thread_tag = f"thr{ALGORITHM_CONFIG.get('THREADS', 0)}"
    fname = f"lbbd_main_{DATASET_NAME}_{sample_tag}_{thread_tag}_{ts}.xlsx"
    return str(Path(OUTPUT_DIR) / fname)


def dataset_prefix_for_name(dataset_name: str) -> str:
    """Return PSPLIB filename prefix, e.g. J30 -> j30, J60 -> j60, J90 -> j90."""
    ds = str(dataset_name).strip().lower()
    if ds.startswith("j"):
        return ds
    return "j" + ds


def instance_name(group: int, sample_index: int) -> str:
    prefix = dataset_prefix_for_name(DATASET_NAME)
    return f"{prefix}{group}_{sample_index}.sm"


def find_instance(root: str, group: int, sample_index: int) -> Optional[str]:
    name = instance_name(group, sample_index)
    p = Path(root) / name
    if p.exists():
        return str(p)
    matches = list(Path(root).glob(f"**/{name}"))
    return str(matches[0]) if matches else None


def build_cases() -> List[RunCase]:
    cases: List[RunCase] = []
    run_id = 0
    for gamma in GAMMA_LIST:
        for alpha in ALPHA_LIST:
            for tl in TIME_LIMIT_LIST:
                for group in range(GROUP_START, GROUP_END + 1):
                    for idx in SAMPLE_INDICES:
                        fp = find_instance(DATASET_ROOT, group, idx)
                        if fp is None:
                            if SKIP_MISSING_FILES:
                                print(f"[WARN] missing file: {instance_name(group, idx)}", flush=True)
                                continue
                            fp = str(Path(DATASET_ROOT) / instance_name(group, idx))
                        run_id += 1
                        cases.append(RunCase(run_id, DATASET_NAME, group, idx, int(gamma), float(alpha), float(tl), fp))
    return cases


def apply_algorithm_config(case: RunCase) -> None:
    algorithm.FILE_PATH = case.file_path
    algorithm.GAMMA = int(case.gamma)
    algorithm.ALPHA = float(case.alpha)
    algorithm.TIME_LIMIT = float(case.time_limit)
    for k, val in ALGORITHM_CONFIG.items():
        setattr(algorithm, k, copy.deepcopy(val))


def calc_gap(obj: Optional[float], bound: Optional[float]) -> Optional[float]:
    if obj is None or bound is None:
        return None
    try:
        return max(0.0, (float(obj) - float(bound)) / max(1.0, abs(float(obj))))
    except Exception:
        return None


def format_float(v: Any, digits: int = 4) -> str:
    try:
        x = float(v)
        if math.isinf(x) or math.isnan(x):
            return "-"
        if abs(x) >= 1e5 or (abs(x) > 0 and abs(x) < 1e-3):
            return f"{x:.2e}"
        return f"{x:.{digits}f}"
    except Exception:
        return "-"


def live_line(case: RunCase, idx: int, total: int, elapsed: float, best: Any, bound: Any, gap: Any, nodes: Any, stats: Dict[str, Any]) -> str:
    """Return a compact progress line without newline or carriage return.

    The caller decides whether to overwrite the current terminal line. The line is
    intentionally short to avoid wrapping in PowerShell, because wrapping is the
    main reason a carriage-return progress display becomes many lines.
    """
    file_short = Path(case.file_path).name
    if len(file_short) > 16:
        file_short = file_short[:13] + "..."
    gap_pct = "-" if gap is None else f"{100*float(gap):5.1f}%"
    return (
        f"[{idx:03d}/{total:03d}] {file_short:<16} "
        f"G={case.gamma:<3} a={case.alpha:<4} "
        f"t={elapsed:6.1f}s best={format_float(best):>7} bd={format_float(bound):>7} "
        f"gap={gap_pct:>7} node={float(nodes or 0):7.0f} "
        f"SP={stats.get('mipsol',0):3} MFS={stats.get('mfs',0):3} PCu={stats.get('path_user',0):3}"
    )


def terminal_progress_write(model: gp.Model, text: str) -> None:
    """Print live progress as one overwritten terminal line.

    Uses carriage return + right padding to erase leftovers from the previous
    update. Also truncates to terminal width so PowerShell does not wrap it into
    multiple lines.
    """
    if LIVE_PROGRESS_MODE.lower() == "newline":
        print(text, flush=True)
        return
    width = LIVE_PROGRESS_MAX_WIDTH
    if not width or width <= 0:
        try:
            width = shutil.get_terminal_size((120, 20)).columns
        except Exception:
            width = 120
    # Leave a small margin to avoid terminal wrapping at the exact boundary.
    width = max(40, int(width) - 2)
    if len(text) > width:
        text = text[:max(0, width - 3)] + "..."
    prev_len = int(getattr(model, "_last_live_line_len", 0))
    pad = max(0, prev_len - len(text))
    sys.stdout.write("\r" + text + (" " * pad))
    sys.stdout.flush()
    model._last_live_line_len = len(text)


def terminal_progress_clear(model: gp.Model) -> None:
    if LIVE_PROGRESS_MODE.lower() == "newline":
        return
    prev_len = int(getattr(model, "_last_live_line_len", 0))
    if prev_len > 0:
        sys.stdout.write("\r" + (" " * prev_len) + "\r")
        sys.stdout.flush()
        model._last_live_line_len = 0




def result_for_popup(result: Dict[str, Any], elapsed: Optional[float] = None, case: Optional[RunCase] = None) -> Dict[str, Any]:
    """Convert LBBD result dict to the generic popup dashboard schema."""
    d = dict(result)
    if case is not None:
        d.setdefault("method", "lbbd_main")
        d.setdefault("time_limit", case.time_limit)
        d.setdefault("gamma", case.gamma)
        d.setdefault("alpha", case.alpha)
        d.setdefault("file", case.file_path)
    else:
        d.setdefault("method", "lbbd_main")
    if elapsed is not None:
        d["elapsed"] = elapsed
    return d

def solve_one(case: RunCase, writer: ExcelLiveWriter, case_index: int, total_cases: int, dashboard=None) -> Dict[str, Any]:
    apply_algorithm_config(case)
    start_wall = time.time()
    start_time_str = now_str()

    result: Dict[str, Any] = {
        "run_id": case.run_id,
        "dataset": case.dataset,
        "group": case.group,
        "sample_index": case.sample_index,
        "gamma": case.gamma,
        "alpha": case.alpha,
        "time_limit": case.time_limit,
        "file": case.file_path,
        "status": None,
        "status_name": "RUNNING",
        "abnormal": False,
        "error": "",
        "objective": None,
        "best_bound": None,
        "gap": None,
        "runtime_solver": 0.0,
        "runtime_total": 0.0,
        "build_time": 0.0,
        "nodes": 0,
        "sol_count": 0,
        "vars": 0,
        "constrs": 0,
        "pair_cuts": 0,
        "initial_lb": None,
        "initial_ub": None,
        "initial_rule": None,
        "sp_calls": 0,
        "sp_infeasible": 0,
        "sp_feasible": 0,
        "mfs_cuts": 0,
        "mfs_repeat": 0,
        "mfs_miss": 0,
        "mfs_avg_size": None,
        "nogood_cuts": 0,
        "cache_size": 0,
        "cache_hits": 0,
        "path_pool": 0,
        "path_user": 0,
        "path_lazy": 0,
        "path_cert": 0,
        "path_cert_fail": 0,
        "verify_trob": None,
        "verify_resource_feasible": None,
        "start_time": start_time_str,
        "end_time": None,
        "last_update": start_time_str,
    }
    writer.add_or_update_result(result, force_save=True)
    if dashboard is not None:
        dashboard.draw(case_index - 1, case_index, result_for_popup(result, 0.0, case), force=True)

    try:
        data = algorithm.load_psplib_sm(case.file_path)
        initial_lb = algorithm.robust_cpm_lb_original(data, int(case.gamma), float(case.alpha))
        sgs = algorithm.serial_sgs_initial_solution(data, int(case.gamma), float(case.alpha)) if algorithm.USE_INITIAL_SGS_UB else {"rule": "none", "sigma": None, "UB": float("inf")}
        initial_ub = float(sgs.get("UB", float("inf")))
        initial_sigma = sgs.get("sigma") if isinstance(sgs.get("sigma"), dict) else None
        result.update({
            "initial_lb": initial_lb,
            "initial_ub": None if math.isinf(initial_ub) else initial_ub,
            "initial_rule": sgs.get("rule", "none"),
        })
        writer.add_or_update_result(result, force_save=False)

        build_t0 = time.time()
        model, sigma, eta = algorithm.build_hybrid_model(data, initial_lb, initial_ub, initial_sigma)
        build_time = time.time() - build_t0

        model.Params.LazyConstraints = 1
        model.Params.PreCrush = 1
        model._data = data
        model._sigma_vars = sigma
        model._eta_var = eta
        model._gamma = int(case.gamma)
        model._alpha = float(case.alpha)
        model._initial_lb = initial_lb
        model._use_path_cuts = algorithm.ADD_PATH_OPT_CUTS
        model._sp_cache = {}
        model._lazy_signatures = set()
        model._path_cut_pool = []
        model._path_cut_pool_sigs = set()
        model._path_user_sigs = set()
        model._stat = {
            "mipsol": 0, "inf": 0, "feas": 0,
            "mfs": 0, "mfs_repeat": 0, "mfs_miss": 0, "mfs_size_sum": 0,
            "nogood": 0, "cache_hit": 0,
            "path_cert": 0, "path_cert_fail": 0, "path_pool": 0, "path_pool_repeat": 0,
            "path_len_sum": 0, "path_user": 0, "path_lazy": 0, "path_pool_scans": 0,
        }
        model._callback_error = None
        model._progress_output = False
        model._progress_interval = LIVE_PRINT_INTERVAL_SEC
        model._last_live_print_time = 0.0
        model._last_live_line_len = 0
        model._last_excel_save_time = 0.0
        model._solve_start_time = time.time()

        result.update({"build_time": build_time, "vars": model.NumVars, "constrs": model.NumConstrs, "pair_cuts": getattr(model, "_pair_cnt", 0)})
        writer.add_or_update_result(result, force_save=False)

        def wrapper_callback(m: gp.Model, where: int) -> None:
            # Live progress on MIP callback. Delegate MIPNODE/MIPSOL to the algorithm.
            if where == GRB.Callback.MIP:
                now = time.time()
                if now - getattr(m, "_last_live_print_time", 0.0) >= LIVE_PRINT_INTERVAL_SEC:
                    try:
                        elapsed = now - m._solve_start_time
                        node = m.cbGet(GRB.Callback.MIP_NODCNT)
                        best = m.cbGet(GRB.Callback.MIP_OBJBST)
                        bound = m.cbGet(GRB.Callback.MIP_OBJBND)
                        solcnt = m.cbGet(GRB.Callback.MIP_SOLCNT)
                        gap = None
                        if solcnt > 0 and abs(best) < 1e99 and abs(bound) < 1e99:
                            gap = calc_gap(best, bound)
                        if dashboard is not None and UI_MODE.lower() in ("popup", "external", "gui", "window"):
                            popup_update = result_for_popup(result, elapsed, case)
                            popup_update.update({
                                "objective": None if solcnt <= 0 or abs(best) > 1e99 else float(best),
                                "best_bound": None if abs(bound) > 1e99 else float(bound),
                                "gap": gap,
                                "nodes": float(node),
                                "sol_count": int(solcnt),
                                "status_name": "RUNNING",
                                "sp_calls": int(m._stat.get("mipsol", 0)),
                                "mfs_cuts": int(m._stat.get("mfs", 0)),
                                "path_user": int(m._stat.get("path_user", 0)),
                            })
                            dashboard.draw(case_index - 1, case_index, popup_update, force=False)
                        else:
                            terminal_progress_write(m, live_line(case, case_index, total_cases, elapsed, best, bound, gap, node, m._stat))
                        m._last_live_print_time = now

                        if SAVE_EXCEL_DURING_SOLVE and now - getattr(m, "_last_excel_save_time", 0.0) >= EXCEL_SAVE_INTERVAL_SEC:
                            result.update({
                                "objective": None if solcnt <= 0 or abs(best) > 1e99 else float(best),
                                "best_bound": None if abs(bound) > 1e99 else float(bound),
                                "gap": gap,
                                "runtime_solver": elapsed,
                                "runtime_total": time.time() - start_wall,
                                "nodes": float(node),
                                "sol_count": int(solcnt),
                                "last_update": now_str(),
                                "sp_calls": int(m._stat.get("mipsol", 0)),
                                "sp_infeasible": int(m._stat.get("inf", 0)),
                                "sp_feasible": int(m._stat.get("feas", 0)),
                                "mfs_cuts": int(m._stat.get("mfs", 0)),
                                "mfs_repeat": int(m._stat.get("mfs_repeat", 0)),
                                "mfs_miss": int(m._stat.get("mfs_miss", 0)),
                                "nogood_cuts": int(m._stat.get("nogood", 0)),
                                "cache_size": len(getattr(m, "_sp_cache", {})),
                                "cache_hits": int(m._stat.get("cache_hit", 0)),
                                "path_pool": int(m._stat.get("path_pool", 0)),
                                "path_user": int(m._stat.get("path_user", 0)),
                                "path_lazy": int(m._stat.get("path_lazy", 0)),
                                "path_cert": int(m._stat.get("path_cert", 0)),
                                "path_cert_fail": int(m._stat.get("path_cert_fail", 0)),
                            })
                            writer.add_or_update_result(result, force_save=False)
                            m._last_excel_save_time = now
                    except Exception:
                        pass
                return
            algorithm._hybrid_callback(m, where)

        model.optimize(wrapper_callback)
        if CLEAR_LIVE_LINE_AFTER_SOLVE and not (dashboard is not None and UI_MODE.lower() in ("popup", "external", "gui", "window")):
            terminal_progress_clear(model)
        else:
            sys.stdout.write("\n")
            sys.stdout.flush()

        runtime_total = time.time() - start_wall
        status = int(model.Status)
        status_name = STATUS_NAME.get(status, str(status))
        sol_count = int(model.SolCount)
        obj = float(model.ObjVal) if sol_count > 0 else None
        bound = float(model.ObjBound) if hasattr(model, "ObjBound") else None
        gap = calc_gap(obj, bound) if obj is not None and bound is not None else None
        stats = model._stat
        verify_trob = None
        verify_res = None
        if VERIFY_INCUMBENT and sol_count > 0:
            sigma_sol = {a: 1 if var.X > 0.5 else 0 for a, var in sigma.items()}
            try:
                verify_trob = algorithm.robust_longest_path_value(data, sigma_sol, int(case.gamma), float(case.alpha))
            except Exception:
                verify_trob = "verify_error"
            try:
                verify_res = bool(algorithm.check_resource_flow_closure(data, sigma_sol).get("feasible", False))
            except Exception:
                verify_res = False

        mfs_avg = None
        if int(stats.get("mfs", 0)) > 0:
            mfs_avg = float(stats.get("mfs_size_sum", 0)) / max(1, int(stats.get("mfs", 0)))

        abnormal = (status != GRB.OPTIMAL) or (gap is None and status != GRB.OPTIMAL) or (gap is not None and gap > ABNORMAL_GAP_TOL)
        if verify_res is False:
            abnormal = True
        if getattr(model, "_callback_error", None):
            abnormal = True

        result.update({
            "status": status,
            "status_name": status_name,
            "abnormal": bool(abnormal),
            "error": getattr(model, "_callback_error", None) or "",
            "objective": obj,
            "best_bound": bound,
            "gap": gap,
            "runtime_solver": float(getattr(model, "Runtime", runtime_total)),
            "runtime_total": runtime_total,
            "build_time": build_time,
            "nodes": float(getattr(model, "NodeCount", 0.0)),
            "sol_count": sol_count,
            "vars": int(model.NumVars),
            "constrs": int(model.NumConstrs),
            "pair_cuts": int(getattr(model, "_pair_cnt", 0)),
            "sp_calls": int(stats.get("mipsol", 0)),
            "sp_infeasible": int(stats.get("inf", 0)),
            "sp_feasible": int(stats.get("feas", 0)),
            "mfs_cuts": int(stats.get("mfs", 0)),
            "mfs_repeat": int(stats.get("mfs_repeat", 0)),
            "mfs_miss": int(stats.get("mfs_miss", 0)),
            "mfs_avg_size": mfs_avg,
            "nogood_cuts": int(stats.get("nogood", 0)),
            "cache_size": len(getattr(model, "_sp_cache", {})),
            "cache_hits": int(stats.get("cache_hit", 0)),
            "path_pool": int(stats.get("path_pool", 0)),
            "path_user": int(stats.get("path_user", 0)),
            "path_lazy": int(stats.get("path_lazy", 0)),
            "path_cert": int(stats.get("path_cert", 0)),
            "path_cert_fail": int(stats.get("path_cert_fail", 0)),
            "verify_trob": verify_trob,
            "verify_resource_feasible": verify_res,
            "end_time": now_str(),
            "last_update": now_str(),
        })
        writer.add_or_update_result(result, force_save=True)
        writer.write_abnormal_if_needed(result)
        if dashboard is not None:
            popup_final = result_for_popup(result, runtime_total, case)
            dashboard.add_completed(popup_final)
            dashboard.draw(case_index, case_index, popup_final, force=True)

        if PRINT_FINAL_ONE_LINE and not (UI_MODE.lower() in ("popup", "external", "gui", "window")):
            print(
                f"[DONE] {Path(case.file_path).name} G={case.gamma} a={case.alpha} "
                f"status={status_name} obj={format_float(obj)} bd={format_float(bound)} "
                f"gap={'-' if gap is None else f'{100*gap:.3f}%'} time={runtime_total:.2f}s "
                f"SP={result['sp_calls']} MFS={result['mfs_cuts']} PCuser={result['path_user']} "
                f"abnormal={result['abnormal']}",
                flush=True,
            )
        return result

    except Exception as exc:
        sys.stdout.write("\n")
        err = repr(exc)
        tb = traceback.format_exc(limit=5)
        result.update({
            "status": None,
            "status_name": "ERROR",
            "abnormal": True,
            "error": err + "\n" + tb,
            "runtime_total": time.time() - start_wall,
            "end_time": now_str(),
            "last_update": now_str(),
        })
        writer.add_or_update_result(result, force_save=True)
        writer.write_abnormal_if_needed(result)
        if dashboard is not None:
            popup_final = result_for_popup(result, result.get("runtime_total"), case)
            dashboard.add_completed(popup_final)
            dashboard.draw(case_index, case_index, popup_final, force=True)
        print(f"[ERROR] {Path(case.file_path).name}: {err}", flush=True)
        if STOP_AFTER_FIRST_ERROR:
            raise
        return result


def print_run_header(cases: List[RunCase], output_path: str) -> None:
    print("========== Robust RCPSP LBBD-main benchmark library ==========")
    print(f"dataset       : {DATASET_NAME}")
    print(f"root          : {DATASET_ROOT}")
    print(f"groups        : {GROUP_START}..{GROUP_END}")
    print(f"sample indices: {SAMPLE_INDICES}")
    print(f"Gamma list    : {GAMMA_LIST}")
    print(f"alpha list    : {ALPHA_LIST}")
    print(f"time limits   : {TIME_LIMIT_LIST}")
    print(f"threads       : {ALGORITHM_CONFIG.get('THREADS', 0)}")
    print(f"total cases   : {len(cases)}")
    print(f"output xlsx   : {output_path}")
    print("progress      : one overwritten live line; workbook is saved during run")
    print("=============================================================")


def main() -> None:
    cases = build_cases()
    output_path = make_output_path()
    writer = ExcelLiveWriter(output_path)
    run_info = RunInfo(
        name="LBBD-main robust RCPSP library",
        method="lbbd_main",
        dataset=DATASET_NAME,
        output_path=output_path,
        total_jobs=len(cases),
        time_limit=max(TIME_LIMIT_LIST) if TIME_LIMIT_LIST else 0,
        threads=ALGORITHM_CONFIG.get("THREADS", 0),
        gamma_list=GAMMA_LIST,
        alpha_list=ALPHA_LIST,
        group_range=f"{GROUP_START}..{GROUP_END}",
        sample_indices=SAMPLE_INDICES,
    )
    dashboard = create_dashboard(run_info, mode=UI_MODE, refresh_interval=POPUP_REFRESH_INTERVAL_SEC, enabled=(UI_MODE.lower() != "off"), launch_popup=LAUNCH_POPUP_UI)
    print_run_header(cases, output_path)
    if not cases:
        print("No cases found. Check DATASET_ROOT / GROUP range / SAMPLE_INDICES.")
        return

    all_start = time.time()
    for idx, case in enumerate(cases, start=1):
        print(
            f"\n[START] {idx}/{len(cases)} file={Path(case.file_path).name} "
            f"Gamma={case.gamma} alpha={case.alpha} TL={case.time_limit}s",
            flush=True,
        )
        solve_one(case, writer, idx, len(cases), dashboard=dashboard)
        writer.rebuild_summary()

    writer.rebuild_summary()
    try:
        dashboard.draw(len(cases), len(cases), None, force=True)
    except Exception:
        pass
    print("\n========== Benchmark finished ==========")
    print(f"total wall time: {time.time() - all_start:.2f}s")
    print(f"Excel saved    : {output_path}")
    print("========================================")


if __name__ == "__main__":
    main()
